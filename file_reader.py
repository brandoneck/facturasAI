def read_txt(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def read_docx(path):
    from docx import Document
    doc = Document(path)
    return "\n".join([p.text for p in doc.paragraphs])


def read_excel(path):
    from openpyxl import load_workbook
    wb = load_workbook(path)
    sheet = wb.active
    return "\n".join([str(row) for row in sheet.iter_rows(values_only=True)])


def read_file(path):
    if path.endswith(".txt"):
        return read_txt(path)
    elif path.endswith(".docx"):
        return read_docx(path)
    elif path.endswith(".xlsx"):
        return read_excel(path)
    else:
        raise ValueError("Formato no soportado")
